"""Tests for B-164 -- FCA short-interest ingest (Sprint 61).

Covers (plan section 2, step 7):
  1. Migration 014: chain head pin "14"; short_positions + isin_ticker_map
     tables exist; both indexes exist.
  2. Name normaliser cases.
  3. Excel datetime -> ISO date conversion.
  4. Sheet-prefix matching (FCA dates the sheet names daily).
  5. Workbook parse from an in-test 2-sheet XLSX fixture (no network):
     0.0 exit rows KEPT, incomplete rows skipped, --from/--to filter.
  6. Upsert idempotency: run-twice leaves the row count unchanged;
     a changed pct updates in place; mapped ticker survives re-upsert.
  7. As-of aggregate semantics: per-holder latest row wins, updates
     supersede, 0.0 exits self-cancel, None when no data at/before date.
  8. ISIN->ticker mapping: unambiguous name match, ambiguity dropped,
     manual override CSV wins (all with use_figi=False -- no network).
  9. Coverage calculation on a synthetic universe.
 10. backtest HEADER: 64 columns, short_pct_at_announcement directly
     before the Sprint 63 pairs (B-155/B-159/B-161) and
     windows_available.

ASCII-only prints; tempfile DBs; no network anywhere.
Requires openpyxl (also a runtime dependency of the ingest script).
"""
from __future__ import annotations

import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest import mock

HERE = Path(__file__).resolve().parent
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))

import backfill_short_interest as bsi  # noqa: E402


# ---------------------------------------------------------------------------
# Fixture builders
# ---------------------------------------------------------------------------

FIXTURE_ROWS_CURRENT = [
    # (holder, issuer, isin, pct, position_date)
    ("Kintbury Capital LLP", "Aberdeen Group plc",
     "GB00BF8Q6K64", 1.44, datetime(2026, 4, 15)),
    ("MECM, Limited", "Aberdeen Group plc",
     "GB00BF8Q6K64", 0.51, datetime(2025, 4, 30)),
    ("SurgoCap Partners LP", "Wynnstay Group PLC",
     "GB0034212331", 0.62, datetime(2026, 5, 20)),
]

FIXTURE_ROWS_HISTORIC = [
    ("AlphaGen Capital Limited", "1SPATIAL PLC",
     "GB00B09LQS34", 0, datetime(2013, 6, 24)),          # int 0 exit row -- KEPT
    ("AlphaGen Capital Limited", "1SPATIAL PLC",
     "GB00B09LQS34", 1.42, datetime(2013, 5, 29)),
    ("AlphaGen Capital Limited", "1SPATIAL PLC",
     "GB00B09LQS34", 0.71, datetime(2013, 5, 24)),
    # Incomplete rows -- skipped by the parser:
    ("", "No Holder plc", "GB00MISSING1", 0.9, datetime(2020, 1, 1)),
    ("No ISIN Fund LP", "Mystery plc", "", 0.9, datetime(2020, 1, 1)),
    ("No Date Fund LP", "Dateless plc", "GB00NODATE11", 0.9, None),
]


def build_fixture_xlsx(path: Path,
                       current_name: str = "Current Disclosures 10.06.2026",
                       historic_name: str = "Historic Disclosures 10.06.2026",
                       ) -> Path:
    """Write the 2-sheet fixture workbook mirroring the live FCA layout."""
    import openpyxl

    wb = openpyxl.Workbook()
    header = ("Position Holder", "Name of Share Issuer", "ISIN",
              "Net Short Position (%)", "Position Date")
    ws_cur = wb.active
    ws_cur.title = current_name
    ws_cur.append(header)
    for row in FIXTURE_ROWS_CURRENT:
        ws_cur.append(row)
    ws_his = wb.create_sheet(historic_name)
    ws_his.append(header)
    for row in FIXTURE_ROWS_HISTORIC:
        ws_his.append(row)
    wb.save(str(path))
    return path


def _insert_tx(conn, fingerprint, ticker, company, *, tx_type="BUY",
               announced_at="", tx_date="2026-05-01"):
    conn.execute(
        "INSERT INTO transactions (fingerprint, first_seen, last_seen, "
        "seen_count, date, ticker, company, director, type, shares, price, "
        "value, announced_at) "
        "VALUES (?, '2026-01-01T00:00:00Z', '2026-01-01T00:00:00Z', 1, ?, "
        "?, ?, 'Jane Doe', ?, 100, 1.0, 100.0, ?)",
        (fingerprint, tx_date, ticker, company, tx_type, announced_at),
    )


def _insert_sp(conn, holder, isin, pct, pdate, *, ticker=None,
               issuer="Issuer plc"):
    conn.execute(
        "INSERT INTO short_positions (position_holder, issuer_name, isin, "
        "ticker, net_short_pct, position_date, source, fetched_at) "
        "VALUES (?, ?, ?, ?, ?, ?, 'ssr_daily', '2026-06-10T00:00:00Z')",
        (holder, issuer, isin, ticker, pct, pdate),
    )


class TempDBTestCase(unittest.TestCase):
    """Shared tempfile-DB scaffolding (FUSE-safe: tempdir, never .data/)."""

    def setUp(self):
        import db as db_mod
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        db_path = Path(self._tmp.name) / "test.db"
        patcher = mock.patch.object(db_mod, "DB_PATH", db_path)
        patcher.start()
        self.addCleanup(patcher.stop)
        self.conn = db_mod.connect()
        self.addCleanup(self.conn.close)
        self.db_mod = db_mod
        # Guard: never read a real .data/_isin_overrides.csv during tests.
        ov_patch = mock.patch.object(
            bsi, "OVERRIDES_CSV", Path(self._tmp.name) / "no_overrides.csv")
        ov_patch.start()
        self.addCleanup(ov_patch.stop)


# ---------------------------------------------------------------------------
# 1. Migration 014
# ---------------------------------------------------------------------------

class TestMigration014(TempDBTestCase):

    def test_schema_head_is_14(self):
        self.assertEqual(
            self.db_mod.get_meta(self.conn, "schema_version"), "14")

    def test_tables_exist(self):
        names = {r[0] for r in self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")}
        self.assertIn("short_positions", names)
        self.assertIn("isin_ticker_map", names)

    def test_indexes_exist(self):
        names = {r[0] for r in self.conn.execute(
            "SELECT name FROM sqlite_master WHERE type='index'")}
        self.assertIn("idx_short_positions_ticker_date", names)
        self.assertIn("idx_short_positions_isin", names)

    def test_position_holder_nullable_for_ansp(self):
        # Future ANSP rows are anonymised -- holder NULL must be accepted.
        self.conn.execute(
            "INSERT INTO short_positions (position_holder, issuer_name, "
            "isin, net_short_pct, position_date, source, fetched_at) "
            "VALUES (NULL, 'Issuer plc', 'GB00TESTANSP', 2.1, '2026-08-01', "
            "'ansp_monthly', '2026-08-01T00:00:00Z')")
        n = self.conn.execute(
            "SELECT COUNT(*) FROM short_positions").fetchone()[0]
        self.assertEqual(n, 1)


# ---------------------------------------------------------------------------
# 2. Name normaliser
# ---------------------------------------------------------------------------

class TestNameNormaliser(unittest.TestCase):

    def test_cases(self):
        cases = [
            ("Aberdeen Group plc", "ABERDEEN"),
            ("ABERDEEN GROUP PLC.", "ABERDEEN"),
            ("Wynnstay Group PLC", "WYNNSTAY"),
            ("1SPATIAL PLC", "1SPATIAL"),
            ("SIG plc", "SIG"),
            ("Marks & Spencer Group plc", "MARKS SPENCER"),
            ("J.D. Wetherspoon PLC", "J D WETHERSPOON"),
            ("Foo Holdings Limited", "FOO"),
            ("Bar Ltd", "BAR"),
            ("  spaced   out  plc ", "SPACED OUT"),
            ("", ""),
        ]
        for raw, expected in cases:
            self.assertEqual(
                bsi.normalise_company_name(raw), expected, f"input: {raw!r}")

    def test_suffix_only_name_normalises_empty(self):
        self.assertEqual(bsi.normalise_company_name("Group Holdings PLC"), "")


# ---------------------------------------------------------------------------
# 3. Excel datetime -> ISO conversion
# ---------------------------------------------------------------------------

class TestToIsoDate(unittest.TestCase):

    def test_datetime(self):
        self.assertEqual(
            bsi.to_iso_date(datetime(2026, 4, 15, 0, 0)), "2026-04-15")

    def test_date(self):
        self.assertEqual(bsi.to_iso_date(date(2013, 6, 24)), "2013-06-24")

    def test_iso_string(self):
        self.assertEqual(bsi.to_iso_date("2026-04-15"), "2026-04-15")
        self.assertEqual(bsi.to_iso_date("2026-04-15T00:00:00"), "2026-04-15")

    def test_uk_string(self):
        self.assertEqual(bsi.to_iso_date("15/04/2026"), "2026-04-15")
        self.assertEqual(bsi.to_iso_date("15.04.2026"), "2026-04-15")

    def test_garbage(self):
        self.assertIsNone(bsi.to_iso_date(None))
        self.assertIsNone(bsi.to_iso_date(""))
        self.assertIsNone(bsi.to_iso_date("not a date"))
        self.assertIsNone(bsi.to_iso_date(42.5))


# ---------------------------------------------------------------------------
# 4. Sheet-prefix matching
# ---------------------------------------------------------------------------

class TestSheetMatching(unittest.TestCase):

    def test_prefix_match_with_dated_names(self):
        names = ["Current Disclosures 10.06.2026",
                 "Historic Disclosures 10.06.2026"]
        self.assertEqual(
            bsi.match_sheet_name(names, "Current Disclosures"), names[0])
        self.assertEqual(
            bsi.match_sheet_name(names, "Historic Disclosures"), names[1])

    def test_case_insensitive(self):
        names = ["CURRENT DISCLOSURES 01.01.2027"]
        self.assertEqual(
            bsi.match_sheet_name(names, "current disclosures"), names[0])

    def test_no_match_returns_none(self):
        self.assertIsNone(
            bsi.match_sheet_name(["Sheet1", "Notes"], "Current Disclosures"))


# ---------------------------------------------------------------------------
# 5. Workbook parse (in-test XLSX fixture, no network)
# ---------------------------------------------------------------------------

class TestParseWorkbook(unittest.TestCase):

    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)
        self.xlsx = build_fixture_xlsx(Path(self._tmp.name) / "fixture.xlsx")

    def test_parses_both_sheets(self):
        parsed = bsi.parse_workbook(self.xlsx)
        self.assertEqual(parsed["sheets"]["current"], 3)
        # 6 historic fixture rows, 3 incomplete -> skipped.
        self.assertEqual(parsed["sheets"]["historic"], 3)
        self.assertEqual(len(parsed["rows"]), 6)

    def test_zero_pct_exit_row_kept(self):
        parsed = bsi.parse_workbook(self.xlsx)
        zeros = [r for r in parsed["rows"] if r["net_short_pct"] == 0.0]
        self.assertEqual(len(zeros), 1)
        self.assertEqual(zeros[0]["isin"], "GB00B09LQS34")
        self.assertEqual(zeros[0]["position_date"], "2013-06-24")

    def test_excel_datetime_converted(self):
        parsed = bsi.parse_workbook(self.xlsx)
        dates = {r["position_date"] for r in parsed["rows"]}
        self.assertIn("2026-04-15", dates)
        self.assertIn("2013-05-24", dates)
        for d in dates:
            self.assertRegex(d, r"^\d{4}-\d{2}-\d{2}$")

    def test_from_to_filter(self):
        parsed = bsi.parse_workbook(
            self.xlsx, date_from="2026-01-01", date_to="2026-12-31")
        self.assertEqual(len(parsed["rows"]), 2)   # the two 2026 rows
        self.assertEqual(parsed["skipped_filter"], 4)

    def test_missing_sheet_warns_not_raises(self):
        path = build_fixture_xlsx(
            Path(self._tmp.name) / "oneSheet.xlsx",
            historic_name="Something Else Entirely")
        parsed = bsi.parse_workbook(path)
        self.assertEqual(parsed["sheets"]["historic"], 0)
        self.assertEqual(parsed["sheets"]["current"], 3)


# ---------------------------------------------------------------------------
# 6. Upsert idempotency
# ---------------------------------------------------------------------------

class TestUpsertIdempotency(TempDBTestCase):

    def _rows(self, pct=1.44):
        return [{
            "position_holder": "Kintbury Capital LLP",
            "issuer_name": "Aberdeen Group plc",
            "isin": "GB00BF8Q6K64",
            "net_short_pct": pct,
            "position_date": "2026-04-15",
            "sheet": "current",
        }]

    def test_run_twice_no_duplicates(self):
        s1 = bsi.upsert_short_positions(
            self.conn, self._rows(), "2026-06-10T00:00:00Z")
        s2 = bsi.upsert_short_positions(
            self.conn, self._rows(), "2026-06-11T00:00:00Z")
        self.assertEqual(s1["inserted"], 1)
        self.assertEqual(s2["inserted"], 0)
        self.assertEqual(s2["table_total"], 1)

    def test_conflict_updates_pct_in_place(self):
        bsi.upsert_short_positions(
            self.conn, self._rows(1.44), "2026-06-10T00:00:00Z")
        bsi.upsert_short_positions(
            self.conn, self._rows(1.60), "2026-06-11T00:00:00Z")
        row = self.conn.execute(
            "SELECT net_short_pct, fetched_at FROM short_positions").fetchone()
        self.assertAlmostEqual(row[0], 1.60)
        self.assertEqual(row[1], "2026-06-11T00:00:00Z")

    def test_mapped_ticker_survives_reupsert(self):
        bsi.upsert_short_positions(
            self.conn, self._rows(), "2026-06-10T00:00:00Z")
        self.conn.execute(
            "UPDATE short_positions SET ticker = 'ABDN'")
        bsi.upsert_short_positions(
            self.conn, self._rows(1.50), "2026-06-11T00:00:00Z")
        row = self.conn.execute(
            "SELECT ticker, net_short_pct FROM short_positions").fetchone()
        self.assertEqual(row[0], "ABDN")
        self.assertAlmostEqual(row[1], 1.50)


# ---------------------------------------------------------------------------
# 7. As-of aggregate semantics
# ---------------------------------------------------------------------------

class TestAggregateShortPct(TempDBTestCase):

    def test_latest_per_holder_wins_and_sums(self):
        _insert_sp(self.conn, "Fund A", "GB1", 1.2, "2026-01-10", ticker="TST")
        _insert_sp(self.conn, "Fund A", "GB1", 1.5, "2026-02-01", ticker="TST")
        _insert_sp(self.conn, "Fund B", "GB1", 0.7, "2026-01-20", ticker="TST")
        # As of 2026-02-15: A's latest is 1.5, B's is 0.7 -> 2.2.
        self.assertAlmostEqual(
            bsi.aggregate_short_pct(self.conn, "TST", "2026-02-15"), 2.2)
        # As of 2026-01-15: A's latest is 1.2, B not yet disclosed -> 1.2.
        self.assertAlmostEqual(
            bsi.aggregate_short_pct(self.conn, "TST", "2026-01-15"), 1.2)

    def test_exit_rows_self_cancel(self):
        _insert_sp(self.conn, "Fund A", "GB1", 1.5, "2026-01-10", ticker="TST")
        _insert_sp(self.conn, "Fund A", "GB1", 0.0, "2026-02-01", ticker="TST")
        _insert_sp(self.conn, "Fund B", "GB1", 0.6, "2026-01-20", ticker="TST")
        # After A exits, only B's 0.6 remains.
        self.assertAlmostEqual(
            bsi.aggregate_short_pct(self.conn, "TST", "2026-02-10"), 0.6)

    def test_all_exited_returns_zero_not_none(self):
        _insert_sp(self.conn, "Fund A", "GB1", 1.5, "2026-01-10", ticker="TST")
        _insert_sp(self.conn, "Fund A", "GB1", 0.0, "2026-02-01", ticker="TST")
        self.assertEqual(
            bsi.aggregate_short_pct(self.conn, "TST", "2026-03-01"), 0.0)

    def test_no_data_returns_none(self):
        _insert_sp(self.conn, "Fund A", "GB1", 1.5, "2026-01-10", ticker="TST")
        # Before any disclosure -> None ("no data" != "0.0% short").
        self.assertIsNone(
            bsi.aggregate_short_pct(self.conn, "TST", "2025-12-31"))
        # Unknown ticker -> None.
        self.assertIsNone(
            bsi.aggregate_short_pct(self.conn, "OTHER", "2026-02-01"))
        # Missing args -> None.
        self.assertIsNone(bsi.aggregate_short_pct(self.conn, "TST", ""))
        self.assertIsNone(bsi.aggregate_short_pct(self.conn, "", "2026-02-01"))

    def test_datetime_string_truncated_to_date(self):
        _insert_sp(self.conn, "Fund A", "GB1", 1.5, "2026-01-10", ticker="TST")
        self.assertAlmostEqual(
            bsi.aggregate_short_pct(
                self.conn, "TST", "2026-01-10T16:45:08Z"), 1.5)

    def test_inclusive_false_excludes_same_day_position(self):
        # Lookahead-bias guard (P3-6): the FCA publishes a position the
        # NEXT business day, so a position dated ON the announcement
        # date was not public knowledge at announcement time. The
        # backtest column path calls with inclusive=False and must NOT
        # see it; the default helper (inclusive=True) still does.
        _insert_sp(self.conn, "Fund A", "GB1", 1.5, "2026-01-10", ticker="TST")
        # Default (as-of) semantics: same-day row included.
        self.assertAlmostEqual(
            bsi.aggregate_short_pct(self.conn, "TST", "2026-01-10"), 1.5)
        # Strictly-prior semantics: same-day row excluded -> no data.
        self.assertIsNone(
            bsi.aggregate_short_pct(self.conn, "TST", "2026-01-10",
                                    inclusive=False))

    def test_inclusive_false_still_sees_prior_days(self):
        _insert_sp(self.conn, "Fund A", "GB1", 1.2, "2026-01-09", ticker="TST")
        _insert_sp(self.conn, "Fund A", "GB1", 1.5, "2026-01-10", ticker="TST")
        # inclusive=False on the 10th: only the 09th row is knowable.
        self.assertAlmostEqual(
            bsi.aggregate_short_pct(self.conn, "TST", "2026-01-10",
                                    inclusive=False), 1.2)
        # inclusive=True picks up the same-day update.
        self.assertAlmostEqual(
            bsi.aggregate_short_pct(self.conn, "TST", "2026-01-10"), 1.5)


# ---------------------------------------------------------------------------
# 8. ISIN -> ticker mapping (no network: use_figi=False)
# ---------------------------------------------------------------------------

class TestMapping(TempDBTestCase):

    def test_unambiguous_name_match(self):
        _insert_tx(self.conn, "fp1", "ABDN", "Aberdeen Group plc")
        _insert_sp(self.conn, "Fund A", "GB00BF8Q6K64", 1.4, "2026-04-15",
                   issuer="ABERDEEN GROUP PLC")
        stats = bsi.map_isins(self.conn, use_figi=False)
        self.assertEqual(stats["by_method"].get("name_match"), 1)
        row = self.conn.execute(
            "SELECT ticker FROM short_positions "
            "WHERE isin = 'GB00BF8Q6K64'").fetchone()
        self.assertEqual(row[0], "ABDN")
        m = self.conn.execute(
            "SELECT ticker, method FROM isin_ticker_map "
            "WHERE isin = 'GB00BF8Q6K64'").fetchone()
        self.assertEqual((m[0], m[1]), ("ABDN", "name_match"))

    def test_ambiguous_name_not_matched(self):
        # Same normalised company name -> two tickers (dual share lines).
        _insert_tx(self.conn, "fp1", "AAA", "Doppel plc")
        _insert_tx(self.conn, "fp2", "BBB", "Doppel Limited")
        _insert_sp(self.conn, "Fund A", "GB00DOPPEL01", 0.8, "2026-01-01",
                   issuer="Doppel plc")
        stats = bsi.map_isins(self.conn, use_figi=False)
        self.assertEqual(stats["by_method"], {})
        row = self.conn.execute(
            "SELECT ticker FROM short_positions").fetchone()
        self.assertIsNone(row[0])

    def test_manual_override_wins(self):
        _insert_tx(self.conn, "fp1", "ABDN", "Aberdeen Group plc")
        _insert_sp(self.conn, "Fund A", "GB00BF8Q6K64", 1.4, "2026-04-15",
                   issuer="Aberdeen Group plc")
        csv_path = Path(self._tmp.name) / "_isin_overrides.csv"
        csv_path.write_text("isin,ticker\nGB00BF8Q6K64,OVRD\n",
                            encoding="utf-8")
        with mock.patch.object(bsi, "OVERRIDES_CSV", csv_path):
            stats = bsi.map_isins(self.conn, use_figi=False)
        self.assertEqual(stats["by_method"].get("manual"), 1)
        row = self.conn.execute(
            "SELECT ticker FROM short_positions").fetchone()
        self.assertEqual(row[0], "OVRD")

    def test_existing_map_skipped_unless_remap(self):
        _insert_tx(self.conn, "fp1", "ABDN", "Aberdeen Group plc")
        _insert_sp(self.conn, "Fund A", "GB00BF8Q6K64", 1.4, "2026-04-15",
                   issuer="Aberdeen Group plc")
        self.conn.execute(
            "INSERT INTO isin_ticker_map (isin, ticker, method, mapped_at) "
            "VALUES ('GB00BF8Q6K64', 'OLD', 'manual', '2026-01-01T00:00:00Z')")
        stats = bsi.map_isins(self.conn, use_figi=False)
        self.assertEqual(stats["newly_resolved"], 0)
        row = self.conn.execute(
            "SELECT ticker FROM short_positions").fetchone()
        self.assertEqual(row[0], "OLD")        # existing map applied to rows
        stats = bsi.map_isins(self.conn, use_figi=False, remap=True)
        self.assertEqual(stats["by_method"].get("name_match"), 1)
        row = self.conn.execute(
            "SELECT ticker FROM short_positions").fetchone()
        self.assertEqual(row[0], "ABDN")       # remap re-resolves by name


# ---------------------------------------------------------------------------
# 9. Coverage calculation
# ---------------------------------------------------------------------------

class TestCoverage(unittest.TestCase):

    def test_synthetic_universe(self):
        buys = {
            "AAA": ["2026-01-15"],                  # disclosure 30d away -> covered
            "BBB": ["2026-01-15"],                  # disclosure 200d away -> not
            "CCC": ["2026-01-15", "2025-06-01"],    # second buy is close -> covered
            "DDD": ["2026-01-15"],                  # no disclosures -> not
        }
        discl = {
            "AAA": ["2026-02-14"],
            "BBB": ["2026-08-03"],
            "CCC": ["2025-05-15"],
            "ZZZ": ["2026-01-15"],   # not a BUY ticker -- ignored
        }
        stats = bsi.coverage_pct(buys, discl, window_days=90)
        self.assertEqual(stats["buy_tickers_total"], 4)
        self.assertEqual(stats["buy_tickers_covered"], 2)
        self.assertAlmostEqual(stats["coverage_pct"], 50.0)

    def test_exact_window_boundary_inclusive(self):
        buys = {"AAA": ["2026-01-01"]}
        discl = {"AAA": ["2026-04-01"]}   # exactly 90 days
        stats = bsi.coverage_pct(buys, discl, window_days=90)
        self.assertEqual(stats["buy_tickers_covered"], 1)

    def test_empty_universe(self):
        stats = bsi.coverage_pct({}, {}, window_days=90)
        self.assertEqual(stats["buy_tickers_total"], 0)
        self.assertEqual(stats["coverage_pct"], 0.0)


class TestCoverageReportDB(TempDBTestCase):
    """coverage_report end-to-end on a tempfile DB + tempfile JSONL."""

    def test_report_and_jsonl_append(self):
        import json as _json
        _insert_tx(self.conn, "fp1", "AAA", "Aaa plc",
                   announced_at="2026-01-15T07:00:00Z")
        _insert_tx(self.conn, "fp2", "BBB", "Bbb plc",
                   announced_at="2025-01-15T07:00:00Z")
        _insert_sp(self.conn, "Fund A", "GBAAA", 1.0, "2026-02-01",
                   ticker="AAA")
        jsonl = Path(self._tmp.name) / "_short_coverage.jsonl"
        stats = bsi.coverage_report(self.conn, jsonl_path=jsonl)
        self.assertEqual(stats["buy_tickers_total"], 2)
        self.assertEqual(stats["buy_tickers_covered"], 1)
        self.assertAlmostEqual(stats["coverage_pct"], 50.0)
        lines = jsonl.read_text(encoding="utf-8").strip().splitlines()
        self.assertEqual(len(lines), 1)
        logged = _json.loads(lines[0])
        self.assertEqual(logged["buy_tickers_covered"], 1)
        self.assertEqual(logged["rows_total"], 1)


# ---------------------------------------------------------------------------
# 10. backtest HEADER integration
# ---------------------------------------------------------------------------

class TestBacktestHeader(unittest.TestCase):

    def test_header_has_64_columns(self):
        import backtest as bt
        # 57 (post-B-156) + short_pct_at_announcement = 58;
        # + routine_flag + routine_prior_buy_years (B-155) = 60;
        # + seller_reversal_flag + net_shares_prior_12m (B-159) = 62;
        # + post_results_flag + days_since_results (B-161) = 64.
        self.assertEqual(len(bt.HEADER), 64)

    def test_short_pct_before_sprint63_pairs_and_windows_available(self):
        import backtest as bt
        # B-155, B-159 then B-161 (Sprint 63) each inserted a pair
        # directly before windows_available, shifting short_pct back by 6.
        idx_wa = bt.HEADER.index("windows_available")
        self.assertEqual(bt.HEADER[idx_wa - 1], "days_since_results")
        self.assertEqual(bt.HEADER[idx_wa - 2], "post_results_flag")
        self.assertEqual(bt.HEADER[idx_wa - 3], "net_shares_prior_12m")
        self.assertEqual(bt.HEADER[idx_wa - 4], "seller_reversal_flag")
        self.assertEqual(bt.HEADER[idx_wa - 5], "routine_prior_buy_years")
        self.assertEqual(bt.HEADER[idx_wa - 6], "routine_flag")
        self.assertEqual(bt.HEADER[idx_wa - 7], "short_pct_at_announcement")
        self.assertEqual(bt.HEADER[idx_wa - 8], "holding_pct_increase")

    def test_backtest_imports_canonical_aggregate(self):
        import backtest as bt
        # The canonical implementation is the one in the ingest script
        # (try/except ImportError fallback keeps backtest alive without it).
        self.assertIs(bt.aggregate_short_pct, bsi.aggregate_short_pct)

    def test_backtest_call_is_strictly_prior(self):
        # Lookahead-bias guard (P3-6): the backtest feature column must
        # use strictly-prior data (inclusive=False) because the FCA
        # publishes positions the next business day. Source-level check
        # so a regression to the default (inclusive) call fails loudly.
        import backtest as bt
        src = Path(bt.__file__).read_text(encoding="utf-8")
        call_pos = src.index("aggregate_short_pct(conn, ticker, announced")
        snippet = src[call_pos:call_pos + 200]
        self.assertIn("inclusive=False", snippet)


if __name__ == "__main__":
    unittest.main()

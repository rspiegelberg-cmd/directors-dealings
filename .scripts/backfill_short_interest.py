"""backfill_short_interest.py -- FCA short-interest ingest (B-164).

ZONE B -- writes to `.data/directors.db`, `.scripts/_short_cache/` and
`.data/_short_coverage.jsonl`. **Rupert runs this**; Claude never runs it
from bash.

What it does
------------
Downloads the FCA's Short Selling Regulation daily disclosure workbook
(net short positions >= 0.5% of issued share capital, 0.1% increments,
published each UK business day), parses BOTH sheets, and writes the rows
into the `short_positions` table. ISINs are mapped to our tickers via:

  (a) normalised company-name match against distinct (company, ticker)
      pairs from `transactions` (strip PLC/LIMITED/LTD/GROUP/HOLDINGS,
      punctuation, case; unambiguous matches only);
  (b) OpenFIGI ISIN lookup fallback, ONLY for unmapped ISINs whose
      normalised issuer name fuzzy-overlaps our universe (skippable with
      --no-figi -- it needs network);
  (c) manual override CSV `.data/_isin_overrides.csv` (columns: isin,ticker)
      -- highest priority, read if present.

Resolved mappings persist in `isin_ticker_map` so re-runs are cheap.

Source workbook (stable URL, overwritten daily, ~3.1 MB):
    https://www.fca.org.uk/publication/data/short-positions-daily-update.xlsx

Two sheets with DATED names -- matched by prefix:
    "Current Disclosures DD.MM.YYYY"   (~600 rows)
    "Historic Disclosures DD.MM.YYYY"  (~107k rows, back to ~2012)
Columns (verified live 2026-06-10):
    Position Holder | Name of Share Issuer | ISIN |
    Net Short Position (%) | Position Date (Excel datetime)
0.0 rows mean the holder dropped below the 0.5% threshold (an exit) and
are KEPT -- the as-of aggregate needs them so closed positions self-cancel.

REGIME CHANGE 13 Jul 2026 (PS26/5): individual holder-level disclosure
ends; replaced by anonymised aggregated per-issuer data (ANSP). The
historic sheet is a one-shot backfill of ~13 years of holder-level data
that becomes unfetchable after that date -- ingest before then. Dated
copies under `.scripts/_short_cache/` are our own archive.

Coverage report (mandatory, end of every non-dry run): % of distinct
BUY-transaction tickers with >= 1 mapped disclosure within +/-90 days of
any of their BUY announced dates; printed and appended as a JSON line to
`.data/_short_coverage.jsonl`. Expected only ~10-25% -- the 0.5%
disclosure threshold skews large-cap while our universe skews AIM.

Dependency: openpyxl (Rupert: `pip install openpyxl`). Imported lazily so
`aggregate_short_pct` stays importable (backtest.py) without it.

Run:
    python .scripts\\backfill_short_interest.py --dry-run   # parse + mapping report, no writes
    python .scripts\\backfill_short_interest.py             # full ingest + coverage log
    python .scripts\\backfill_short_interest.py --remap     # re-resolve all ISIN mappings
    python .scripts\\backfill_short_interest.py --no-figi   # skip the OpenFIGI network fallback
    python .scripts\\backfill_short_interest.py --from 2024-01-01 --to 2024-12-31
    python .scripts\\backfill_short_interest.py --file path\\to\\archived.xlsx
"""
from __future__ import annotations

import argparse
import csv
import difflib
import json
import re
import sys
import time
from datetime import date, datetime, timezone
from pathlib import Path

HERE = Path(__file__).resolve().parent
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))

import db         # noqa: E402
import db_health  # noqa: E402

# --- Constants --------------------------------------------------------------

FCA_URL = ("https://www.fca.org.uk/publication/data/"
           "short-positions-daily-update.xlsx")
SOURCE = "ssr_daily"
CACHE_DIR = HERE / "_short_cache"
OVERRIDES_CSV = db.DB_DIR / "_isin_overrides.csv"
COVERAGE_JSONL = db.DB_DIR / "_short_coverage.jsonl"
USER_AGENT = (
    "DirectorsDealingsBot/1.0 (personal research project; "
    "contact rspiegelberg@gmail.com)"
)

CURRENT_SHEET_PREFIX = "Current Disclosures"
HISTORIC_SHEET_PREFIX = "Historic Disclosures"

# OpenFIGI free tier without an API key: max 10 mapping jobs per request,
# 25 requests per minute. Stay well inside that.
OPENFIGI_URL = "https://api.openfigi.com/v3/mapping"
OPENFIGI_BATCH = 10
OPENFIGI_DELAY_SECONDS = 2.5
# Gate for sending an unmapped ISIN to OpenFIGI: its normalised issuer
# name must fuzzy-overlap a normalised universe company name at this
# similarity or better (difflib ratio). Keeps us from burning quota on
# the ~90% of disclosed issuers we will never hold.
FIGI_FUZZY_CUTOFF = 0.75

COVERAGE_WINDOW_DAYS = 90

# Corporate-suffix tokens stripped by the name normaliser. "PLC." etc are
# handled because punctuation is stripped first.
_SUFFIX_TOKENS = frozenset({"PLC", "LIMITED", "LTD", "GROUP", "HOLDINGS"})

_NON_ALNUM_RE = re.compile(r"[^A-Z0-9 ]+")


# --- Pure helpers (no network / no DB -- unit-tested) ------------------------

def normalise_company_name(name: str) -> str:
    """Canonical form for issuer-name matching.

    Upper-case, punctuation -> space, strip corporate-suffix tokens
    (PLC / LIMITED / LTD / GROUP / HOLDINGS), collapse whitespace.
    'Aberdeen Group plc' and 'ABERDEEN GROUP PLC.' both -> 'ABERDEEN'.
    """
    s = _NON_ALNUM_RE.sub(" ", (name or "").upper())
    tokens = [t for t in s.split() if t not in _SUFFIX_TOKENS]
    return " ".join(tokens)


def to_iso_date(value) -> str | None:
    """Excel cell value -> 'YYYY-MM-DD', or None when unparseable.

    openpyxl (data_only) hands back datetime objects for date-formatted
    cells; strings appear in hand-edited archive copies.
    """
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        # ISO first (with or without a time part), then UK d/m/Y.
        try:
            return date.fromisoformat(s[:10]).isoformat()
        except ValueError:
            pass
        for fmt in ("%d/%m/%Y", "%d.%m.%Y", "%d-%b-%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def match_sheet_name(sheetnames: list[str], prefix: str) -> str | None:
    """Find the sheet whose name starts with `prefix` (case-insensitive).

    The FCA dates the sheet names ('Current Disclosures 10.06.2026') so an
    exact-name lookup breaks daily; prefix matching does not.
    """
    p = prefix.strip().lower()
    for name in sheetnames:
        if (name or "").strip().lower().startswith(p):
            return name
    return None


def _header_indexes(row: tuple) -> dict | None:
    """Map a candidate header row to column indexes, or None if not a header.

    Matches by prefix so minor FCA renames ('Net Short Position (%)' vs
    'Net Short Position') keep working.
    """
    wanted = {
        "holder": "position holder",
        "issuer": "name of share issuer",
        "isin": "isin",
        "pct": "net short position",
        "pdate": "position date",
    }
    idx: dict = {}
    for i, cell in enumerate(row):
        text = str(cell or "").strip().lower()
        if not text:
            continue
        for key, prefix in wanted.items():
            if key not in idx and text.startswith(prefix):
                idx[key] = i
    return idx if len(idx) == len(wanted) else None


def parse_sheet(ws, sheet_label: str) -> list[dict]:
    """Parse one disclosure sheet into row dicts.

    Returns [{position_holder, issuer_name, isin, net_short_pct,
    position_date, sheet}]. Rows missing holder / ISIN / date / pct are
    skipped (counted by the caller via len()); pct == 0.0 is KEPT.
    """
    rows: list[dict] = []
    idx: dict | None = None
    max_idx = 0
    for row in ws.iter_rows(values_only=True):
        if idx is None:
            idx = _header_indexes(row)
            if idx is not None:
                max_idx = max(idx.values())
            continue
        if len(row) <= max_idx:   # ragged trailing row -- skip
            continue
        holder = str(row[idx["holder"]] or "").strip()
        issuer = str(row[idx["issuer"]] or "").strip()
        isin = str(row[idx["isin"]] or "").strip().upper()
        pct_raw = row[idx["pct"]]
        pdate = to_iso_date(row[idx["pdate"]])
        if not holder or not isin or pdate is None or pct_raw is None:
            continue
        try:
            pct = float(pct_raw)
        except (TypeError, ValueError):
            continue
        rows.append({
            "position_holder": holder,
            "issuer_name": issuer,
            "isin": isin,
            "net_short_pct": pct,
            "position_date": pdate,
            "sheet": sheet_label,
        })
    return rows


def parse_workbook(path: Path, *, date_from: str | None = None,
                   date_to: str | None = None) -> dict:
    """Parse both disclosure sheets from the workbook at `path`.

    Returns {"rows": [...], "sheets": {label: count}, "skipped_filter": n}.
    --from/--to filter on position_date (ISO, inclusive).
    """
    import openpyxl  # lazy -- keeps aggregate_short_pct importable without it

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        all_rows: list[dict] = []
        sheet_counts: dict = {}
        for prefix, label in ((CURRENT_SHEET_PREFIX, "current"),
                              (HISTORIC_SHEET_PREFIX, "historic")):
            name = match_sheet_name(wb.sheetnames, prefix)
            if name is None:
                print(f"[short_interest] WARNING: no sheet matching "
                      f"prefix '{prefix}' in {path.name}")
                sheet_counts[label] = 0
                continue
            rows = parse_sheet(wb[name], label)
            sheet_counts[label] = len(rows)
            all_rows.extend(rows)
    finally:
        wb.close()

    n_before = len(all_rows)
    if date_from:
        all_rows = [r for r in all_rows if r["position_date"] >= date_from]
    if date_to:
        all_rows = [r for r in all_rows if r["position_date"] <= date_to]
    return {
        "rows": all_rows,
        "sheets": sheet_counts,
        "skipped_filter": n_before - len(all_rows),
    }


def coverage_pct(buy_dates_by_ticker: dict, disclosure_dates_by_ticker: dict,
                 window_days: int = COVERAGE_WINDOW_DAYS) -> dict:
    """Pure coverage calculation (unit-tested without a DB).

    buy_dates_by_ticker: {ticker: [iso buy dates]}
    disclosure_dates_by_ticker: {ticker: [iso position dates]} (mapped only)
    A ticker is covered when ANY disclosure falls within +/-window_days of
    ANY of its BUY dates.
    """
    covered = set()
    for ticker, buy_dates in buy_dates_by_ticker.items():
        ddates = disclosure_dates_by_ticker.get(ticker)
        if not ddates:
            continue
        for b in buy_dates:
            try:
                bd = date.fromisoformat(str(b)[:10])
            except (TypeError, ValueError):
                continue
            for d in ddates:
                try:
                    dd = date.fromisoformat(str(d)[:10])
                except (TypeError, ValueError):
                    continue
                if abs((dd - bd).days) <= window_days:
                    covered.add(ticker)
                    break
            if ticker in covered:
                break
    total = len(buy_dates_by_ticker)
    return {
        "buy_tickers_total": total,
        "buy_tickers_covered": len(covered),
        "coverage_pct": round(100.0 * len(covered) / total, 1) if total else 0.0,
    }


# --- Canonical aggregate (imported by backtest.py) ---------------------------

def aggregate_short_pct(conn, ticker: str, on_date: str,
                        inclusive: bool = True):
    """Aggregate net short interest in `ticker` as of `on_date`.

    As-of join: for each position holder, take their most recent
    disclosure with position_date <= on_date, then sum the percentages.
    Exit rows (0.0) therefore self-cancel: once a holder's latest row is
    0.0 they contribute nothing. Implemented with a ROW_NUMBER() window
    CTE (SQLite >= 3.25).

    `inclusive=False` makes the comparison strictly prior
    (position_date < on_date). The FCA publishes a position the NEXT
    business day, so a position dated ON an announcement date was not
    yet public knowledge at announcement time -- backtest.py uses
    inclusive=False to avoid lookahead bias (P3-6). The default (True)
    preserves the general as-of helper semantics.

    Returns a float (sum of holder pcts, possibly 0.0 when every holder
    has exited) or None when there is no disclosure data at or before
    `on_date` -- callers must keep "no data" distinct from "0.0% short".
    `on_date` may carry a time part; only the first 10 chars are used.
    """
    if not ticker or not on_date:
        return None
    cmp_op = "<=" if inclusive else "<"
    row = conn.execute(
        f"""
        WITH latest AS (
            SELECT net_short_pct,
                   ROW_NUMBER() OVER (
                       PARTITION BY position_holder
                       ORDER BY position_date DESC, id DESC
                   ) AS rn
            FROM short_positions
            WHERE ticker = ?
              AND position_date {cmp_op} ?
              AND position_holder IS NOT NULL
        )
        SELECT SUM(net_short_pct) AS total_pct, COUNT(*) AS n
        FROM latest WHERE rn = 1
        """,
        (ticker, str(on_date)[:10]),
    ).fetchone()
    # Access by NAME, not position: psycopg dict_row has no integer keys, and
    # unnamed aggregates get PG-assigned names — so alias them. (B-180)
    if row is None or not row["n"]:
        return None
    return round(float(row["total_pct"]), 4)


# --- Network -----------------------------------------------------------------

def download_workbook(*, use_cache: bool = True) -> Path:
    """Download today's workbook to the dated cache path.

    Skips the download when today's file is already cached (the FCA
    overwrites the same URL daily, so one fetch per day is enough). The
    dated copies double as our own archive ahead of the 13 Jul 2026
    regime change.
    """
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    target = CACHE_DIR / f"short-positions-{date.today():%Y%m%d}.xlsx"
    if use_cache and target.exists() and target.stat().st_size > 0:
        print(f"[short_interest] using cached {target.name} "
              f"({target.stat().st_size} bytes)")
        return target

    import requests  # lazy

    print(f"[short_interest] downloading {FCA_URL}")
    resp = requests.get(FCA_URL, headers={"User-Agent": USER_AGENT},
                        timeout=120)
    resp.raise_for_status()
    target.write_bytes(resp.content)
    print(f"[short_interest] saved {target.name} ({len(resp.content)} bytes)")
    return target


def _openfigi_lookup(isins: list[str]) -> dict:
    """Map ISIN -> exchange ticker via OpenFIGI (free tier, no key).

    Returns {isin: ticker} for resolved ISINs only. Network errors are
    non-fatal: a warning is printed and the remaining ISINs stay unmapped.
    """
    import requests  # lazy

    out: dict = {}
    for i in range(0, len(isins), OPENFIGI_BATCH):
        batch = isins[i:i + OPENFIGI_BATCH]
        jobs = [{"idType": "ID_ISIN", "idValue": isin, "exchCode": "LN"}
                for isin in batch]
        try:
            resp = requests.post(
                OPENFIGI_URL, json=jobs,
                headers={"User-Agent": USER_AGENT}, timeout=30)
            resp.raise_for_status()
            results = resp.json()
        except Exception as exc:  # noqa: BLE001 -- fallback path, never fatal
            print(f"[short_interest] WARNING: OpenFIGI request failed "
                  f"({exc}); skipping remaining lookups")
            return out
        for isin, result in zip(batch, results):
            data = (result or {}).get("data") or []
            for entry in data:
                t = (entry.get("ticker") or "").strip().upper()
                if t:
                    out[isin] = t
                    break
        if i + OPENFIGI_BATCH < len(isins):
            time.sleep(OPENFIGI_DELAY_SECONDS)
    return out


# --- DB write ----------------------------------------------------------------

def upsert_short_positions(conn, rows: list[dict], fetched_at: str) -> dict:
    """Idempotent upsert keyed on UNIQUE(position_holder, isin,
    position_date, source). Re-runs update net_short_pct / issuer_name /
    fetched_at in place; the mapped ticker (if any) is preserved.
    """
    before = conn.execute(
        "SELECT COUNT(*) FROM short_positions").fetchone()[0]
    conn.executemany(
        "INSERT INTO short_positions "
        "(position_holder, issuer_name, isin, net_short_pct, position_date, "
        " source, fetched_at) VALUES (?, ?, ?, ?, ?, ?, ?) "
        "ON CONFLICT(position_holder, isin, position_date, source) "
        "DO UPDATE SET net_short_pct = excluded.net_short_pct, "
        "              issuer_name   = excluded.issuer_name, "
        "              fetched_at    = excluded.fetched_at",
        [(r["position_holder"], r["issuer_name"], r["isin"],
          r["net_short_pct"], r["position_date"], SOURCE, fetched_at)
         for r in rows],
    )
    after = conn.execute(
        "SELECT COUNT(*) FROM short_positions").fetchone()[0]
    return {"rows_in": len(rows), "inserted": after - before,
            "updated_or_unchanged": len(rows) - (after - before),
            "table_total": after}


def build_name_index(conn) -> dict:
    """Normalised company name -> ticker, from distinct transactions pairs.

    Ambiguous names (one normalised name -> several tickers, e.g. dual
    share lines) are dropped -- name-matching must be unambiguous; the
    override CSV handles the awkward cases.
    """
    pairs = conn.execute(
        "SELECT DISTINCT company, ticker FROM transactions "
        "WHERE company IS NOT NULL AND company != '' "
        "  AND ticker IS NOT NULL AND ticker != ''").fetchall()
    index: dict = {}
    ambiguous: set = set()
    for row in pairs:
        norm = normalise_company_name(row[0])
        if not norm:
            continue
        t = str(row[1]).strip().upper()
        if norm in index and index[norm] != t:
            ambiguous.add(norm)
            continue
        index[norm] = t
    for norm in ambiguous:
        index.pop(norm, None)
    return index


def load_overrides(path: Path | None = None) -> dict:
    """Read the manual override CSV (isin,ticker) if present.

    `path` defaults to OVERRIDES_CSV resolved at call time (not bound at
    def time) so tests can patch the module constant.
    """
    if path is None:
        path = OVERRIDES_CSV
    if not path.exists():
        return {}
    out: dict = {}
    with path.open("r", encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            isin = (row.get("isin") or "").strip().upper()
            ticker = (row.get("ticker") or "").strip().upper()
            if isin and ticker:
                out[isin] = ticker
    return out


def map_isins(conn, *, use_figi: bool = True, remap: bool = False,
              dry_run: bool = False) -> dict:
    """Resolve ISIN -> ticker for disclosure rows and persist the mappings.

    Priority: manual override CSV > existing isin_ticker_map (unless
    --remap) > normalised-name match > OpenFIGI (gated by fuzzy name
    overlap with our universe; --no-figi skips it). FIGI results are only
    accepted when the returned ticker is one we actually hold.
    """
    mapped_at = db.iso_now()
    name_index = build_name_index(conn)
    held = {t for t in name_index.values()}
    held |= {str(r[0]).strip().upper() for r in conn.execute(
        "SELECT DISTINCT ticker FROM transactions WHERE ticker != ''")}
    overrides = load_overrides()

    existing: dict = {}
    if not remap:
        existing = {r[0]: r[1] for r in conn.execute(
            "SELECT isin, ticker FROM isin_ticker_map")}

    todo = conn.execute(
        "SELECT isin, MAX(issuer_name) AS issuer_name "
        "FROM short_positions GROUP BY isin").fetchall()

    resolved: dict = {}      # isin -> (ticker, method)
    figi_candidates: list = []   # (isin, norm_issuer)
    universe_names = list(name_index.keys())

    for row in todo:
        isin, issuer = row[0], row[1] or ""
        if isin in overrides:
            resolved[isin] = (overrides[isin], "manual")
            continue
        if isin in existing:
            continue  # already persisted; row-level UPDATE below re-applies
        norm = normalise_company_name(issuer)
        if norm and norm in name_index:
            resolved[isin] = (name_index[norm], "name_match")
            continue
        if use_figi and norm and universe_names:
            close = difflib.get_close_matches(
                norm, universe_names, n=1, cutoff=FIGI_FUZZY_CUTOFF)
            if close:
                figi_candidates.append(isin)

    n_figi_hits = 0
    if use_figi and figi_candidates and not dry_run:
        print(f"[short_interest] OpenFIGI lookups for "
              f"{len(figi_candidates)} fuzzy-gated ISIN(s)")
        figi_map = _openfigi_lookup(figi_candidates)
        for isin, ticker in figi_map.items():
            # Accept only tickers we hold (with a dot-stripped variant
            # tolerance, e.g. FIGI 'BT/A' or 'BT.A' vs our 'BT.A').
            t = ticker.replace("/", ".")
            if t in held:
                resolved[isin] = (t, "openfigi")
                n_figi_hits += 1
            elif t.replace(".", "") in {h.replace(".", "") for h in held}:
                canon = next(h for h in held
                             if h.replace(".", "") == t.replace(".", ""))
                resolved[isin] = (canon, "openfigi")
                n_figi_hits += 1

    if not dry_run:
        conn.executemany(
            "INSERT OR REPLACE INTO isin_ticker_map "
            "(isin, ticker, method, mapped_at) VALUES (?, ?, ?, ?)",
            [(isin, t, m, mapped_at) for isin, (t, m) in resolved.items()],
        )
        # Apply the full persisted map to disclosure rows (covers both the
        # new resolutions and pre-existing map entries on fresh rows).
        conn.execute(
            "UPDATE short_positions SET ticker = "
            "  (SELECT m.ticker FROM isin_ticker_map m "
            "   WHERE m.isin = short_positions.isin) "
            "WHERE isin IN (SELECT isin FROM isin_ticker_map)")

    n_isins = len(todo)
    n_mapped = conn.execute(
        "SELECT COUNT(DISTINCT isin) FROM short_positions "
        "WHERE ticker IS NOT NULL").fetchone()[0] if not dry_run else None
    by_method: dict = {}
    for _isin, (_t, m) in resolved.items():
        by_method[m] = by_method.get(m, 0) + 1
    return {
        "isins_total": n_isins,
        "newly_resolved": len(resolved),
        "by_method": by_method,
        "figi_candidates": len(figi_candidates),
        "figi_hits": n_figi_hits,
        "isins_mapped_in_table": n_mapped,
    }


def coverage_report(conn, *, jsonl_path: Path = COVERAGE_JSONL,
                    write: bool = True,
                    window_days: int = COVERAGE_WINDOW_DAYS) -> dict:
    """Mandatory end-of-run coverage metric.

    % of distinct BUY-transaction tickers with >= 1 mapped disclosure
    within +/-window_days of ANY of their BUY announced dates. Printed and
    (unless write=False) appended as a JSON line to jsonl_path.
    """
    buy_dates: dict = {}
    for row in conn.execute(
        "SELECT ticker, substr(COALESCE(NULLIF(announced_at, ''), date), 1, 10) "
        "FROM transactions WHERE type = 'BUY' AND ticker != ''"
    ):
        buy_dates.setdefault(str(row[0]).strip().upper(), []).append(row[1])

    disclosure_dates: dict = {}
    for row in conn.execute(
        "SELECT ticker, position_date FROM short_positions "
        "WHERE ticker IS NOT NULL"
    ):
        disclosure_dates.setdefault(
            str(row[0]).strip().upper(), []).append(row[1])

    stats = coverage_pct(buy_dates, disclosure_dates, window_days)
    stats.update({
        "ts": db.iso_now(),
        "window_days": window_days,
        "rows_total": conn.execute(
            "SELECT COUNT(*) FROM short_positions").fetchone()[0],
        "rows_mapped": conn.execute(
            "SELECT COUNT(*) FROM short_positions "
            "WHERE ticker IS NOT NULL").fetchone()[0],
        "isins_mapped": conn.execute(
            "SELECT COUNT(*) FROM isin_ticker_map").fetchone()[0],
    })
    print(f"[short_interest] coverage: "
          f"{stats['buy_tickers_covered']}/{stats['buy_tickers_total']} "
          f"BUY tickers ({stats['coverage_pct']}%) have >=1 mapped "
          f"disclosure within +/-{window_days}d of a BUY announcement")
    if write:
        jsonl_path.parent.mkdir(parents=True, exist_ok=True)
        with jsonl_path.open("a", encoding="utf-8") as fh:
            fh.write(json.dumps(stats, sort_keys=True) + "\n")
    return stats


# --- CLI ----------------------------------------------------------------------

def run(*, date_from: str | None = None, date_to: str | None = None,
        dry_run: bool = False, remap: bool = False, use_figi: bool = True,
        use_cache: bool = True, file: str | None = None, conn=None) -> dict:
    """Download (or open --file), parse, upsert, map, report. Returns stats."""
    path = Path(file) if file else download_workbook(use_cache=use_cache)
    parsed = parse_workbook(path, date_from=date_from, date_to=date_to)
    rows = parsed["rows"]
    print(f"[short_interest] parsed current={parsed['sheets'].get('current', 0)} "
          f"historic={parsed['sheets'].get('historic', 0)} "
          f"date_filtered_out={parsed['skipped_filter']} "
          f"to_ingest={len(rows)}")

    own_conn = conn is None
    if own_conn:
        conn = db.connect()
    try:
        if dry_run:
            # Preview the prospective name-match rate; no writes at all.
            name_index = build_name_index(conn)
            overrides = load_overrides()
            isins: dict = {}
            for r in rows:
                isins.setdefault(r["isin"], r["issuer_name"])
            n_name = sum(
                1 for isin, issuer in isins.items()
                if isin not in overrides
                and normalise_company_name(issuer) in name_index)
            n_override = sum(1 for isin in isins if isin in overrides)
            print(f"[short_interest] DRY RUN: {len(isins)} distinct ISINs; "
                  f"name-match would map {n_name}, overrides {n_override}; "
                  f"no DB writes, no coverage log")
            return {"dry_run": True, "rows": len(rows),
                    "isins": len(isins), "name_match": n_name,
                    "overrides": n_override}

        fetched_at = db.iso_now()
        upsert = upsert_short_positions(conn, rows, fetched_at)
        print(f"[short_interest] upsert: rows_in={upsert['rows_in']} "
              f"inserted={upsert['inserted']} "
              f"updated_or_unchanged={upsert['updated_or_unchanged']} "
              f"table_total={upsert['table_total']}")
        mapping = map_isins(conn, use_figi=use_figi, remap=remap)
        print(f"[short_interest] mapping: isins={mapping['isins_total']} "
              f"newly_resolved={mapping['newly_resolved']} "
              f"by_method={mapping['by_method']} "
              f"figi_candidates={mapping['figi_candidates']} "
              f"figi_hits={mapping['figi_hits']} "
              f"isins_mapped_in_table={mapping['isins_mapped_in_table']}")
        conn.commit()
        stats = coverage_report(conn)
        return {"upsert": upsert, "mapping": mapping, "coverage": stats}
    finally:
        if own_conn:
            conn.close()


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        description="Ingest FCA short-interest disclosures (B-164).")
    ap.add_argument("--from", dest="date_from", default=None,
                    help="only ingest rows with position_date >= this "
                         "(ISO YYYY-MM-DD, inclusive).")
    ap.add_argument("--to", dest="date_to", default=None,
                    help="only ingest rows with position_date <= this "
                         "(ISO YYYY-MM-DD, inclusive).")
    ap.add_argument("--dry-run", action="store_true",
                    help="parse + mapping report; write nothing.")
    ap.add_argument("--remap", action="store_true",
                    help="re-resolve every ISIN mapping from scratch "
                         "(ignores existing isin_ticker_map rows).")
    ap.add_argument("--no-figi", action="store_true",
                    help="skip the OpenFIGI network fallback for unmapped "
                         "ISINs (name match + overrides only).")
    ap.add_argument("--no-cache", action="store_true",
                    help="re-download even if today's workbook is cached.")
    ap.add_argument("--file", default=None,
                    help="ingest a local .xlsx (e.g. an archived dated "
                         "copy) instead of downloading.")
    args = ap.parse_args(argv)

    if not args.dry_run:
        if not db_health.check(db.DB_PATH):
            print("[short_interest] FATAL: integrity_check failed.")
            return 2
        if not db_health.backup():
            print("[short_interest] FATAL: backup failed.")
            return 3

    run(date_from=args.date_from, date_to=args.date_to,
        dry_run=args.dry_run, remap=args.remap, use_figi=not args.no_figi,
        use_cache=not args.no_cache, file=args.file)

    if not args.dry_run:
        try:
            if not db_health.check(db.DB_PATH):
                print("[short_interest] WARNING: post-run check failed.")
                return 4
            db_health.seal()
        except Exception as exc:  # noqa: BLE001 -- seal is best-effort
            print("[db_health] seal failed (non-fatal):", exc)
    return 0


if __name__ == "__main__":
    sys.exit(main())
